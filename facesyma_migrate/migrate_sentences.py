"""
migrate_sentences.py
====================
Facesyma MongoDB tam çeviri migration.

Güncellenen veritabanları:
  1. database_attribute_*  — 201 sıfat × 30 karakter cümlesi (Excel'den)
  2. appfaceapi_advisor     — sıfat başına tavsiye cümleleri (cumle_tr → diğer diller)
  3. database_daily_*       — günlük pozitif/negatif cümleler

Güncellenmeyenler (çeviri gerekmez):
  - pos_neg           → sadece sıfat isimleri
  - contrast          → sadece sıfat isimleri
  - appfaceapi_myuser → kullanıcı verisi

KULLANIM:
  pip install pymongo openpyxl deep-translator

  python migrate_sentences.py               # Tüm 7 dil, 3 modül
  python migrate_sentences.py --langs en    # Sadece İngilizce
  python migrate_sentences.py --skip attribute   # Sadece advisor + daily
  python migrate_sentences.py --only daily  # Sadece daily
  python migrate_sentences.py --resume      # Kaldığı yerden devam
  python migrate_sentences.py --dry-run     # Test modu
"""

import os, sys, json, time, argparse, logging
from pathlib  import Path
from datetime import datetime
from pymongo  import MongoClient, UpdateOne
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("migration.log", encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)

# ── Sabitler ─────────────────────────────────────────────────────

MONGO_URI = os.environ.get(
    "MONGO_URI",
    "mongodb+srv://facesyma:FaceSyma2021@cluster0.io98c.mongodb.net/"
    "myFirstDatabase?ssl=true&ssl_cert_reqs=CERT_NONE"
)
EXCEL_PATH = Path(__file__).parent / "Facesyma_201_Sifat_Tam.xlsx"

LANG_MAP = {
    "en": {"attr_db": "database_attribute_en", "daily_db": "database_daily_en",  "gt": "en",  "name": "İngilizce"},
    "de": {"attr_db": "database_attribute_de", "daily_db": "database_daily_de",  "gt": "de",  "name": "Almanca"},
    "ru": {"attr_db": "database_attribute_ru", "daily_db": "database_daily_ru",  "gt": "ru",  "name": "Rusça"},
    "ar": {"attr_db": "database_attribute_ar", "daily_db": "database_daily_ar",  "gt": "ar",  "name": "Arapça"},
    "es": {"attr_db": "database_attribute_sp", "daily_db": None,                  "gt": "es",  "name": "İspanyolca"},
    "ko": {"attr_db": "database_attribute_kr", "daily_db": None,                  "gt": "ko",  "name": "Korece"},
    "ja": {"attr_db": "database_attribute_jp", "daily_db": None,                  "gt": "ja",  "name": "Japonca"},
}

ATTR_COLLECTIONS = ["eye", "eyebrow", "lip", "nose", "forehead"]
BATCH_SIZE       = 10
DELAY_SEC        = 0.6
MAX_RETRY        = 3
PROGRESS_FILE    = Path("migration_progress.json")


# ── Çevirmen ─────────────────────────────────────────────────────

def translate_batch(texts: list, target: str) -> list:
    from deep_translator import GoogleTranslator
    results = []
    for i in range(0, len(texts), BATCH_SIZE):
        chunk = texts[i:i + BATCH_SIZE]
        for attempt in range(1, MAX_RETRY + 1):
            try:
                translated = GoogleTranslator(source="tr", target=target).translate_batch(chunk)
                results.extend(t if t else chunk[j] for j, t in enumerate(translated))
                break
            except Exception as e:
                wait = attempt * 2
                log.warning(f"  Çeviri hatası (deneme {attempt}/{MAX_RETRY}): {e}")
                if attempt < MAX_RETRY:
                    time.sleep(wait)
                else:
                    log.error(f"  Başarısız — orijinal kullanılıyor")
                    results.extend(chunk)
        time.sleep(DELAY_SEC)
    return results


def translate_single(text: str, target: str) -> str:
    r = translate_batch([text], target)
    return r[0] if r else text


# ── İlerleme ─────────────────────────────────────────────────────

def load_progress() -> set:
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return set(json.load(f).get("done", []))
    return set()


def save_progress(done: set):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump({"done": list(done), "ts": datetime.now().isoformat()}, f)


# ── MODÜL 1: database_attribute_* ────────────────────────────────

def load_excel(path: Path) -> dict:
    wb, ws = load_workbook(path), None
    ws     = wb.active
    data, cur = {}, None
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and row[3].value is not None:
            cur = str(row[3].value).strip()
            data[cur] = []
        if row[4].value and row[5].value and cur:
            data[cur].append(str(row[5].value).strip())
    log.info(f"Excel: {len(data)} sıfat, {sum(len(v) for v in data.values()):,} cümle")
    return data


def read_tr_attr_mapping(client: MongoClient) -> dict:
    db, mapping = client["database_attribute_tr"], {}
    for col_name in ATTR_COLLECTIONS:
        for doc in db[col_name].find({}):
            doc_id = doc["_id"]
            for vkey, sdict in doc.items():
                if vkey == "_id" or not isinstance(sdict, dict):
                    continue
                for sifat in sdict:
                    if sifat not in ("att1", "att2"):
                        mapping[sifat] = {"collection": col_name, "doc_id": doc_id, "value_key": vkey}
    log.info(f"TR attribute mapping: {len(mapping)} sıfat")
    return mapping


def migrate_attribute(client, langs, done, dry_run):
    log.info("\n" + "="*55)
    log.info("MODÜL 1: database_attribute_* (201 sıfat × 30 cümle)")
    log.info("="*55)

    excel_data = load_excel(EXCEL_PATH)
    tr_mapping = read_tr_attr_mapping(client)
    common     = sorted(set(excel_data) & set(tr_mapping))
    written    = 0

    for lang_code in langs:
        info     = LANG_MAP[lang_code]
        gt_code  = info["gt"]
        db_target= client[info["attr_db"]]
        log.info(f"\n  [{lang_code.upper()}] {info['name']}")

        bulk = {}
        for idx, sifat in enumerate(common, 1):
            key = f"attr:{lang_code}:{sifat}"
            if key in done:
                continue

            meta      = tr_mapping[sifat]
            path_key  = f"{meta['value_key']}.{sifat}"
            tr_cumles = excel_data[sifat]

            log.info(f"  [{idx:3d}/{len(common)}] {sifat[:45]}")
            translated = translate_batch(tr_cumles, gt_code)

            col_bulk = bulk.setdefault(meta["collection"], {})
            doc_bulk = col_bulk.setdefault(meta["doc_id"], {})
            doc_bulk[path_key] = translated

            done.add(key)
            save_progress(done)
            written += 1

            if written % 20 == 0 and not dry_run:
                _flush_attr(db_target, bulk)
                bulk.clear()
                log.info(f"  → {written} sıfat ara kaydedildi")

        if bulk and not dry_run:
            _flush_attr(db_target, bulk)
        elif dry_run and bulk:
            for cn, docs in bulk.items():
                for did, fields in docs.items():
                    for fp, c in list(fields.items())[:1]:
                        log.info(f"  DRY: {cn}/{did}/{fp} → {c[0][:50]}")
                    break
                break

    return written


def _flush_attr(db, bulk):
    for col_name, docs in bulk.items():
        ops = [UpdateOne({"_id": did}, {"$set": fields}, upsert=True)
               for did, fields in docs.items()]
        if ops:
            r = db[col_name].bulk_write(ops, ordered=False)
            log.debug(f"    flush {col_name}: mod={r.modified_count} ups={r.upserted_count}")


# ── MODÜL 2: appfaceapi_advisor ───────────────────────────────────

def migrate_advisor(client, langs, done, dry_run):
    log.info("\n" + "="*55)
    log.info("MODÜL 2: appfaceapi_advisor (sıfat tavsiye cümleleri)")
    log.info("="*55)

    col     = client["facesyma-backend"]["appfaceapi_advisor"]
    docs    = list(col.find({}))
    written = 0

    log.info(f"  Toplam döküman: {len(docs)}")

    for lang_code in langs:
        info    = LANG_MAP[lang_code]
        gt_code = info["gt"]
        field   = f"cumle_{lang_code}"    # cumle_en, cumle_de, ...
        log.info(f"\n  [{lang_code.upper()}] {info['name']} → alan: {field}")

        ops = []
        for doc in docs:
            sifat = doc.get("sifat", "")
            cumle_tr = doc.get("cumle_tr", "")
            if not cumle_tr:
                continue

            key = f"advisor:{lang_code}:{sifat}"
            if key in done:
                continue

            log.info(f"  {sifat[:40]}")
            translated = translate_single(cumle_tr, gt_code)
            ops.append(UpdateOne(
                {"_id": doc["_id"]},
                {"$set": {field: translated}}
            ))
            done.add(key)
            save_progress(done)
            written += 1

        if ops and not dry_run:
            r = col.bulk_write(ops, ordered=False)
            log.info(f"  {info['name']}: {r.modified_count} döküman güncellendi")
        elif dry_run and ops:
            log.info(f"  DRY: {len(ops)} döküman güncellenecekti")

    return written


# ── MODÜL 3: database_daily_* ─────────────────────────────────────

def migrate_daily(client, langs, done, dry_run):
    log.info("\n" + "="*55)
    log.info("MODÜL 3: database_daily_* (günlük motivasyon cümleleri)")
    log.info("="*55)

    # Sadece daily_db'si olan diller
    daily_langs = [l for l in langs if LANG_MAP[l].get("daily_db")]
    if not daily_langs:
        log.info("  Bu dillerde database_daily koleksiyonu yok, atlanıyor.")
        return 0

    # TR referans
    tr_db   = client["database_daily_tr"]
    tr_pos  = list(tr_db["positive"].find())
    tr_neg  = list(tr_db["negative"].find())
    written = 0

    log.info(f"  TR positive dökümanı: {len(tr_pos)}, negative: {len(tr_neg)}")

    for lang_code in daily_langs:
        info     = LANG_MAP[lang_code]
        gt_code  = info["gt"]
        daily_db = client[info["daily_db"]]
        log.info(f"\n  [{lang_code.upper()}] {info['name']}")

        for col_name, tr_docs, list_key in [
            ("positive", tr_pos, "positive_daily"),
            ("negative", tr_neg, "negative_daily"),
        ]:
            key = f"daily:{lang_code}:{col_name}"
            if key in done:
                log.info(f"  {col_name}: atlandı (tamamlanmış)")
                continue

            for doc in tr_docs:
                tr_cumleler = doc.get(list_key, [])
                if not tr_cumleler:
                    continue

                log.info(f"  {col_name}: {len(tr_cumleler)} cümle çevriliyor...")
                translated = translate_batch(tr_cumleler, gt_code)

                if not dry_run:
                    daily_db[col_name].update_one(
                        {"_id": doc["_id"]},
                        {"$set": {list_key: translated}},
                        upsert=True
                    )
                    log.info(f"  {col_name}: yazıldı ✓")
                else:
                    log.info(f"  DRY: {col_name} → {translated[0][:60]}")

                done.add(key)
                save_progress(done)
                written += len(translated)

    return written


# ── Ana fonksiyon ─────────────────────────────────────────────────

def migrate(langs, modules, dry_run, resume):
    t0     = datetime.now()
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=10000)
    try:
        client.admin.command("ping")
        log.info("MongoDB OK ✓")
    except Exception as e:
        log.error(f"MongoDB bağlantısı başarısız: {e}")
        sys.exit(1)

    done    = load_progress() if resume else set()
    total   = 0

    if "attribute" in modules:
        total += migrate_attribute(client, langs, done, dry_run)

    if "advisor" in modules:
        total += migrate_advisor(client, langs, done, dry_run)

    if "daily" in modules:
        total += migrate_daily(client, langs, done, dry_run)

    elapsed = (datetime.now() - t0).total_seconds()
    log.info(f"\n{'='*55}")
    log.info(f"TAMAMLANDI")
    log.info(f"Süre    : {elapsed:.0f}s ({elapsed/60:.1f} dk)")
    log.info(f"Toplam  : {total:,} cümle/alan işlendi")
    log.info(f"Diller  : {', '.join(langs)}")
    log.info(f"Modüller: {', '.join(modules)}")
    if dry_run:
        log.info("NOT: Dry run — MongoDB'ye yazılmadı.")
    log.info(f"{'='*55}")

    if not dry_run and PROGRESS_FILE.exists():
        PROGRESS_FILE.unlink()


# ── CLI ───────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Facesyma tam veritabanı migration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Modüller:
  attribute  → database_attribute_* (201 sıfat × 30 cümle)
  advisor    → appfaceapi_advisor   (tavsiye cümleleri)
  daily      → database_daily_*    (günlük motivasyon)

Örnekler:
  python migrate_sentences.py                          # Tüm dil + modül
  python migrate_sentences.py --langs en               # Sadece İngilizce
  python migrate_sentences.py --only attribute         # Sadece attribute
  python migrate_sentences.py --skip daily             # Daily hariç
  python migrate_sentences.py --resume                 # Devam ettir
  python migrate_sentences.py --dry-run --langs en     # Test
        """
    )
    p.add_argument("--langs",    default=",".join(LANG_MAP))
    p.add_argument("--only",     default="",                help="Sadece bu modüller (virgülle)")
    p.add_argument("--skip",     default="",                help="Bu modülleri atla (virgülle)")
    p.add_argument("--dry-run",  action="store_true")
    p.add_argument("--resume",   action="store_true")
    p.add_argument("--delay",    type=float, default=DELAY_SEC)
    p.add_argument("--excel",    default=str(EXCEL_PATH))
    args = p.parse_args()

    global DELAY_SEC, EXCEL_PATH
    DELAY_SEC  = args.delay
    EXCEL_PATH = Path(args.excel)

    if not EXCEL_PATH.exists():
        print(f"HATA: Excel bulunamadı → {EXCEL_PATH}")
        sys.exit(1)

    langs = [l.strip() for l in args.langs.split(",") if l.strip() and l.strip() in LANG_MAP]
    if not langs:
        print(f"HATA: Geçerli dil yok. Seçenekler: {list(LANG_MAP.keys())}")
        sys.exit(1)

    all_modules = ["attribute", "advisor", "daily"]
    if args.only:
        modules = [m.strip() for m in args.only.split(",") if m.strip() in all_modules]
    elif args.skip:
        skip    = {m.strip() for m in args.skip.split(",")}
        modules = [m for m in all_modules if m not in skip]
    else:
        modules = all_modules

    # Tahmin
    n_attr    = 201 * 30 * len(langs)
    n_advisor = 201 * len(langs)  # yaklaşık
    n_daily   = 50 * len([l for l in langs if LANG_MAP[l].get("daily_db")]) * 2
    n_batch   = (n_attr + n_advisor + n_daily) / BATCH_SIZE
    est_min   = n_batch * DELAY_SEC / 60

    print(f"""
╔══════════════════════════════════════════════╗
║       Facesyma Tam Veritabanı Migration       ║
╚══════════════════════════════════════════════╝

  Diller  : {', '.join(langs)}
  Modüller: {', '.join(modules)}
  Dry run : {args.dry_run}
  Resume  : {args.resume}

  Tahmini cümle sayısı:
    attribute : ~{n_attr:,}
    advisor   : ~{n_advisor:,}
    daily     : ~{n_daily:,}

  Tahmini süre: ~{est_min:.0f} dakika
""")

    cevap = input("Başlansın mı? (e/h): ").strip().lower()
    if cevap not in ("e", "evet", "y", "yes"):
        print("İptal.")
        sys.exit(0)

    try:
        migrate(langs=langs, modules=modules, dry_run=args.dry_run, resume=args.resume)
    except KeyboardInterrupt:
        log.warning("\nDurduruldu. --resume ile devam edebilirsiniz.")
        sys.exit(0)


if __name__ == "__main__":
    main()
